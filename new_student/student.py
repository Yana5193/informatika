import sqlite3
from openpyxl import load_workbook


connection = sqlite3.connect("st.db")
cursor = connection.cursor()


cursor.execute("""
    CREATE TABLE IF NOT EXISTS level_of_education(
        id_level INTEGER PRIMARY KEY,
        name_level varchar(100) NOT NULL
);
""")

cursor.execute("""
    CREATE TABLE IF NOT EXISTS direction(
        id_direction INTEGER PRIMARY KEY,
        name_direction varchar(100) NOT NULL
);
""")

cursor.execute("""
    CREATE TABLE IF NOT EXISTS types_of_training(
        id_type INTEGER PRIMARY KEY,
        name_type varchar(100) NOT NULL
);
""")

cursor.execute("""
    CREATE TABLE IF NOT EXISTS student(
        id_student INTEGER PRIMARY KEY,
        id_level INTEGER,
        id_direction INTEGER,
        id_type INTEGER,
        surname varchar(30) NOT NULL,
        name varchar(30) NOT NULL,
        patronymic varchar(30),
        average_score INTEGER NOT NULL,
        FOREIGN KEY (id_level) REFERENCES level_of_education(id_level),
        FOREIGN KEY (id_direction) REFERENCES direction(id_direction),
        FOREIGN KEY (id_type) REFERENCES types_of_training(id_type)
);
""")
wb = load_workbook("st.xlsx", read_only=True, data_only=True)
#Заполнение таблиц
# level_of_education
sheet = wb["level_of_education"]
for row in sheet.iter_rows(min_row=2, values_only=True):
    cursor.execute("INSERT OR IGNORE INTO level_of_education (id_level, name_level) VALUES (?, ?)", 
                   (row[0], row[1]))
    
# direction
sheet = wb["direction"]
for row in sheet.iter_rows(min_row=2, values_only=True):
    cursor.execute("INSERT OR IGNORE INTO direction (id_direction, name_direction) VALUES (?, ?)", 
                   (row[0], row[1]))

# types_of_training
sheet = wb["types_of_training"]
for row in sheet.iter_rows(min_row=2, values_only=True):
    if row[0] is None: continue
    cursor.execute("INSERT OR IGNORE INTO types_of_training (id_type, name_type) VALUES (?, ?)", 
                   (row[0], row[1]))

# student
sheet = wb["student"]
for row in sheet.iter_rows(min_row=2, values_only=True):
    if row[0] is None: continue
    cursor.execute("""
        INSERT OR IGNORE INTO student 
        (id_student, id_level, id_direction, id_type, surname, name, patronymic, average_score) 
        VALUES (?, ?, ?, ?, ?, ?, ?, ?)
    """, (row[0], row[1], row[2], row[3], row[4], row[5], row[6], row[7]))

connection.commit()
#количество всех студентов
cursor.execute("""
    SELECT COUNT(*) AS count_students
    FROM student
""")
count_students=cursor.fetchone()[0]
print(f"Всего {count_students} студентов")
print("-"*15)
print("Количество студентов по направлениям:")
cursor.execute("""
    SELECT
        d.name_direction AS direction,
        COUNT(*) AS count_stidents
    FROM
        student s
    JOIN direction d
        ON s.id_direction = d.id_direction
    GROUP BY d.name_direction
""") 
dir=cursor.fetchall()
for i in dir:
    print(f"{i[0]}:{i[1]} чел.")

print("-"*15)
print("Количество студентов по формам обучения (очная, вечерняя, заочная):")
cursor.execute("""
    SELECT
        t.name_type AS types_of_training,
        COUNT(*) AS count_stidents
    FROM
        student s
    JOIN types_of_training t
        ON s.id_type= t.id_type
    GROUP BY t.id_type
""") 
t=cursor.fetchall()
for i in t:
    print(f"{i[0]}:{i[1]} чел.")

print("-"*15)
print("Максимальный, минимальный, средний баллы студентов по направлениям:")
cursor.execute("""
    SELECT
        d.name_direction AS direction,
        MAX(s.average_score) AS max_score,
        MIN(s.average_score) AS max_score,
        ROUND(AVG(s.average_score), 2) AS avg_score
    FROM
        student s
    JOIN direction d
        ON s.id_direction= d.id_direction
    GROUP BY d.name_direction
""") 
r=cursor.fetchall()
for row in r:
    print(f"{row[0]:<25} {row[1]:<8} {row[2]:<8} {row[3]:<8}")

print("-"*15)
print("Средний балл студентов по направления, уровням и формам обучения:")
cursor.execute("""
    SELECT
        d.name_direction AS direction,
        l.name_level AS education_ledel,
        t.name_type AS form,
        ROUND(AVG(s.average_score), 2) AS avg_score     
    FROM
        student s
    JOIN direction d
        ON s.id_direction= d.id_direction
    JOIN level_of_education l
        ON s.id_level=l.id_level
    JOIN types_of_training t
        ON s.id_type=t.id_type
    GROUP BY d.name_direction, l.name_level, t.name_type
""")
p=cursor.fetchall()
for row in p:
    print(f"{row[0]:<25} {row[1]:<15} {row[2]:<15} {row[3]:<15}")

print("-"*15)
print("Для приказа о назначении повышенной стипендии :")
cursor.execute("""
    SELECT
        s.surname,
        s.name,
        s.patronymic,
        s.average_score,
        d.name_direction,
        t.name_type     
    FROM
        student s
    JOIN direction d 
        ON s.id_direction = d.id_direction
    JOIN types_of_training t 
        ON s.id_type = t.id_type
    WHERE d.name_direction = 'прикладная математика'
            AND t.name_type = 'очное'
    ORDER BY s.average_score DESC
    LIMIT 5;
""")
w=cursor.fetchall()
for i, row in enumerate(w, 1):
    print(f"{i:<3} {row[0]:<20} {row[1]:<15} {row[2]:<18} {row[3]:<6} {row[4]}")

print("-"*15)
print("Сколько однофамильцев в данной базе :")
cursor.execute("""
    SELECT COUNT(*) AS students_with_duplicates
FROM student
WHERE surname IN (
    SELECT surname 
    FROM student 
    GROUP BY surname 
    HAVING COUNT(*) > 1
);
""")
result = cursor.fetchone()[0]
print(f"В базе {result} студентов однофамильцев.")
print("-"*15)
print("Полные тезки (совпадают фамилии, имена, отчества):")
cursor.execute("""
SELECT 
    surname,
    name,
    patronymic,
    COUNT(*) AS count
FROM student
GROUP BY surname, name, patronymic
HAVING COUNT(*) > 1
ORDER BY count DESC;
""")
i= cursor.fetchone()
print(f"В базе {i} тезок.")
#ИСПОЛЬЗОВАНИЕ CASE
print("-"*15)
print("Категории студентов по их среднему баллу:")
cursor.execute("""
    SELECT
        CASE
            WHEN average_score =5.00 THEN 'Отличник'
            WHEN average_score >=4.00  THEN 'Хорошист'
            WHEN average_score >= 3.00 THEN 'Троечник'
            ELSE 'Двоечник'
        END as category,
        COUNT(*) AS count_students
    FROM student
    GROUP BY category 
    ORDER BY MIN(average_score) DESC      
""")
c=cursor.fetchall()
for row in c:
    print(f"{row[0]}: {row[1]} чел.")

print("-"*15)
print("Категории студентов по стипендиям:")
cursor.execute("""
    SELECT
        name,
        surname,
        average_score,
               CASE
                    WHEN average_score=5.00 THEN 'Повышенная стипендия'
                    WHEN average_score>=4.00 THEN 'Обычная стипендия'
                    WHEN average_score>=3.00 THEN 'Только социальная стипендия'
                    ELSE 'Нет стипендии'
                END AS stip
    FROM student
 """)
st=cursor.fetchall()
for row in st:
    print(f"{row[0]:<15} {row[1]:<15} {row[3]:<25} {row[2]:>5}")
#ПОДЗАПРОСЫ
print("-"*15)
print("Самый лучший студент в каждом направлении:")
cursor.execute("""
    SELECT 
        d.name_direction,
        s.surname,
        s.name,
        s.average_score
    FROM student s
    JOIN direction d ON s.id_direction = d.id_direction
    WHERE s.average_score = (
        SELECT MAX(s2.average_score)
        FROM student s2
        WHERE s2.id_direction = s.id_direction
    )
    GROUP BY d.name_direction
    ORDER BY d.name_direction
""")
best_students = cursor.fetchall()
for row in best_students:
    print(f"{row[1]} {row[2]} - {row[3]} баллов ({row[0]})")

print("-"*15)
print("Лучший студент в формах образования:")
cursor.execute("""
 SELECT 
        t.name_type,
        s.surname,
        s.name,
        s.average_score
    FROM student s
    JOIN types_of_training t ON s.id_type = t.id_type
    WHERE s.average_score = (
        SELECT MAX(s2.average_score)
        FROM student s2
        WHERE s2.id_type = s.id_type
    )
    ORDER BY t.name_type
""")

best_per_type = cursor.fetchall()
for row in best_per_type:
    print(f"{row[0]}: {row[1]} {row[2]} - {row[3]} баллов")
#CTE
print("-"*15)
print("Количество студентов по уровням обучения :")
cursor.execute("""
    WITH level_count AS (
        SELECT 
            id_level,
            COUNT(*) AS count_students
        FROM student
        GROUP BY id_level
    )
    SELECT 
        l.name_level,
        lc.count_students
    FROM level_count lc
    JOIN level_of_education l ON lc.id_level = l.id_level
    ORDER BY lc.count_students DESC
""")
level_counts = cursor.fetchall()
for row in level_counts:
    print(f"{row[0]}: {row[1]} чел.")

print("-"*15)
print("Средний балл по уровням обучения :")
cursor.execute("""
    WITH level_avg AS (
        SELECT 
            id_level,
            ROUND(AVG(average_score), 2) AS avg_score
        FROM student
        GROUP BY id_level
    )
    SELECT 
        l.name_level,
        la.avg_score
    FROM level_avg la
    JOIN level_of_education l ON la.id_level = l.id_level
    ORDER BY la.avg_score DESC
""")
result = cursor.fetchall()
for row in result:
    print(f"{row[0]}: {row[1]} баллов")
connection.close()