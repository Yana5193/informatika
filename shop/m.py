import sqlite3
from openpyxl import load_workbook
import tkinter as tk
from tkinter import ttk, messagebox
from datetime import datetime

connection = sqlite3.connect("shop.db")
cursor = connection.cursor()

cursor.execute("""
CREATE TABLE IF NOT EXISTS categories (
    id_category INTEGER PRIMARY KEY NOT NULL UNIQUE,
    name_of_category TEXT NOT NULL
);
""")

cursor.execute("""
CREATE TABLE IF NOT EXISTS jobs_titles (
    id INTEGER PRIMARY KEY NOT NULL UNIQUE,
    name TEXT NOT NULL
);
""")

cursor.execute("""
CREATE TABLE IF NOT EXISTS employees (
    id INTEGER PRIMARY KEY NOT NULL UNIQUE,
    name TEXT NOT NULL,
    surname TEXT NOT NULL,
    id_job_tittle INTEGER NOT NULL,
    FOREIGN KEY(id_job_tittle) REFERENCES jobs_titles(id)
);
""")

cursor.execute("""
CREATE TABLE IF NOT EXISTS products (
    id_product INTEGER PRIMARY KEY NOT NULL UNIQUE,
    name_product TEXT NOT NULL,
    price REAL NOT NULL,
    id_category INTEGER NOT NULL,
    quanite_at_strogare REAL NOT NULL,
    FOREIGN KEY(id_category) REFERENCES categories(id_category)
);
""")

cursor.execute("""
CREATE TABLE IF NOT EXISTS receipts (
    id_check INTEGER PRIMARY KEY NOT NULL UNIQUE,
    created_at TEXT NOT NULL,
    id_cashier INTEGER NOT NULL,
    FOREIGN KEY(id_cashier) REFERENCES employees(id)
);
""")

cursor.execute("""
CREATE TABLE IF NOT EXISTS sale_items (
    id_sale INTEGER PRIMARY KEY NOT NULL UNIQUE,
    id_check INTEGER NOT NULL,
    id_product INTEGER NOT NULL,
    quantity REAL NOT NULL,
    FOREIGN KEY(id_check) REFERENCES receipts(id_check),
    FOREIGN KEY(id_product) REFERENCES products(id_product)
);
""")
try:
    mg = load_workbook("C:/Users/anaer/OneDrive/Desktop/учеба/ИГУ/1 курс 2 семестр/информатика/магазин/mg.xlsx", 
                       read_only=True, data_only=True)
    sheet = mg["categories"]
    for row in sheet.iter_rows(min_row=2, values_only=True):
        cursor.execute("INSERT OR IGNORE INTO categories VALUES (?, ?)", row)
    sheet = mg["jobs_titles"]
    for row in sheet.iter_rows(min_row=2, values_only=True):
        cursor.execute("INSERT OR IGNORE INTO jobs_titles VALUES (?, ?)", row)
    sheet = mg["employees"]
    for row in sheet.iter_rows(min_row=2, values_only=True):
        cursor.execute("INSERT OR IGNORE INTO employees VALUES (?, ?, ?, ?)", row)
    sheet = mg["products"]
    for row in sheet.iter_rows(min_row=2, values_only=True):
        cursor.execute("INSERT OR IGNORE INTO products VALUES (?, ?, ?, ?, ?)", row)
    sheet = mg["receipts"]
    for row in sheet.iter_rows(min_row=2, values_only=True):
        cursor.execute("INSERT OR IGNORE INTO receipts VALUES (?, ?, ?)", row)
    sheet = mg["sale_items"]
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if row[0] is None:
            continue
        cursor.execute("INSERT OR IGNORE INTO sale_items VALUES (?, ?, ?, ?)", row)
    connection.commit()
except Exception as e:
    print("Ошибка импорта Excel:", e)

cart = []

def open_new_sale():
    sale_window = tk.Toplevel(root)
    sale_window.title("Новая продажа")
    sale_window.geometry("950x700")
    sale_window.configure(bg="#E6D9FF")

    tk.Label(sale_window, text="Новая продажа", font=("Arial", 18, "bold"), 
             bg="#E6D9FF", fg="#5A3D8C").pack(pady=15)

    columns = ("ID", "Название", "Цена", "Остаток")
    tree = ttk.Treeview(sale_window, columns=columns, show="headings", height=11)
    for col in columns:
        tree.heading(col, text=col)
        tree.column(col, width=190)
    tree.pack(fill="both", expand=True, padx=30, pady=10)

    tk.Label(sale_window, text="Корзина:", font=("Arial", 12, "bold"), 
             bg="#E6D9FF", fg="#333").pack(anchor="w", padx=30, pady=(15,5))
    
    cart_tree = ttk.Treeview(sale_window, columns=("Название", "Количество", "Сумма"), show="headings", height=8)
    for col in ("Название", "Количество", "Сумма"):
        cart_tree.heading(col, text=col)
        cart_tree.column(col, width=220)
    cart_tree.pack(fill="both", expand=True, padx=30, pady=5)

    qty_entry = tk.Entry(sale_window, width=15, font=("Arial", 11))
    qty_entry.pack(pady=10)
    qty_entry.insert(0, "1")

    def load_products():
        tree.delete(*tree.get_children())
        cursor.execute("SELECT id_product, name_product, price, quanite_at_strogare FROM products")
        for row in cursor.fetchall():
            tree.insert("", "end", values=row)

    def add_to_cart():
        selected = tree.focus()
        if not selected:
            messagebox.showwarning("Внимание", "Выберите товар в таблице!")
            return

        values = tree.item(selected, "values")
        product_id = values[0]
        name = values[1]
        price = float(values[2])
        stock = float(values[3])

        try:
            qty = float(qty_entry.get())
            if qty <= 0:
                raise ValueError
        except:
            messagebox.showerror("Ошибка", "Введите корректное количество!")
            return

        if qty > stock:
            messagebox.showerror("Ошибка", "Недостаточно товара на складе!")
            return

        total = price * qty
        cart.append((product_id, price, qty, name))
        cart_tree.insert("", "end", values=(name, qty, f"{total:.2f}"))

        qty_entry.delete(0, tk.END)
        qty_entry.insert(0, "1")
    def remove_from_cart():
        selected = cart_tree.focus() 
        if not selected:
            messagebox.showwarning("Внимание", "Выберите товар в корзине для удаления!")
            return
        item_index = cart_tree.index(selected)
        cart.pop(item_index)
        cart_tree.delete(selected)

    def punch_check():
        if not cart:
            messagebox.showerror("Ошибка", "Корзина пуста!")
            return

        cursor.execute("SELECT id FROM employees LIMIT 1")
        cashier = cursor.fetchone()
        cashier_id = cashier[0] if cashier else 1

        date = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

        cursor.execute("INSERT INTO receipts (created_at, id_cashier) VALUES (?, ?)", 
                       (date, cashier_id))
        check_id = cursor.lastrowid

        total_sum = 0.0
        check_details = []

        for product_id, price, qty, name in cart:
            cursor.execute("INSERT INTO sale_items (id_check, id_product, quantity) VALUES (?, ?, ?)",
                           (check_id, product_id, qty))

            cursor.execute("UPDATE products SET quanite_at_strogare = quanite_at_strogare - ? WHERE id_product = ?",
                           (qty, product_id))

            total_sum += price * qty
            check_details.append(f"{name} — {qty} шт. × {price:.2f} = {price*qty:.2f} руб.")

        connection.commit()

        details_text = "\n".join(check_details)
        messagebox.showinfo("Чек пробит!", 
            f"Чек №{check_id}\n"
            f"Дата: {date}\n\n"
            f"{details_text}\n\n"
            f"ИТОГО: {total_sum:.2f} руб.")
        cart.clear()
        cart_tree.delete(*cart_tree.get_children())
        load_products()

    btn_frame = tk.Frame(sale_window, bg="#E6D9FF")
    btn_frame.pack(pady=20)

    tk.Button(btn_frame, text="Добавить в корзину", font=("Arial", 12), width=22, height=2,
              bg="#B39DFF", fg="white", command=add_to_cart).pack(side="left", padx=20)
    tk.Button(btn_frame,text="Удалить из корзины",font=("Arial",12,"bold"),width=22,height=2,
              bg="#B39DFF",fg="white",command=remove_from_cart).pack(side="left",padx=20)

    tk.Button(btn_frame, text="Пробить чек", font=("Arial", 12, "bold"), width=22, height=2,
              bg="#8A6CFF", fg="white", command=punch_check).pack(side="left", padx=20)

    load_products()  

def show_reports():
    report_window = tk.Toplevel(root)
    report_window.title("Отчёты")
    report_window.geometry("670x580")
    report_window.configure(bg="#E6D9FF")

    tk.Label(report_window, text="Введите дату (YYYY-MM-DD)", 
             font=("Arial", 12), bg="#E6D9FF", fg="#333").pack(pady=15)
    
    date_entry = tk.Entry(report_window, font=("Arial", 12), width=25)
    date_entry.pack()
    date_entry.insert(0, datetime.now().strftime("%Y-%m-%d"))

    result_text = tk.Text(report_window, font=("Consolas", 10), bg="#F5EEFF")
    result_text.pack(fill="both", expand=True, padx=30, pady=15)

    def generate():
        date = date_entry.get().strip()
        result_text.delete("1.0", tk.END)
        cursor.execute("""
            SELECT p.name_product, 
                   SUM(s.quantity) as total_qty, 
                   SUM(p.price * s.quantity) as total_sum
            FROM sale_items s
            JOIN products p ON p.id_product = s.id_product
            JOIN receipts r ON r.id_check = s.id_check
            WHERE DATE(r.created_at) = ?
            GROUP BY p.name_product
            ORDER BY total_sum DESC
        """, (date,))

        rows = cursor.fetchall()

        result_text.insert(tk.END, f"ОТЧЁТ ЗА {date}\n")
        result_text.insert(tk.END, "="*80 + "\n")
        result_text.insert(tk.END, f"{'Товар':<38} {'Кол-во':>12} {'Выручка':>20}\n")
        result_text.insert(tk.END, "-"*80 + "\n")

        for name, qty, total in rows:
            result_text.insert(tk.END, 
                f"{name:<38} {qty:>12.2f} {total or 0:>20.2f} руб.\n")

        cursor.execute("""
            SELECT SUM(p.price * s.quantity)
            FROM sale_items s
            JOIN products p ON p.id_product = s.id_product
            JOIN receipts r ON r.id_check = s.id_check
            WHERE DATE(r.created_at) = ?
        """, (date,))

        revenue = cursor.fetchone()[0] or 0.0

        result_text.insert(tk.END, "-"*80 + "\n")
        result_text.insert(tk.END, f"ОБЩАЯ ВЫРУЧКА ЗА ДЕНЬ: {revenue:.2f} руб.\n")

    tk.Button(report_window, text="Сформировать отчёт", font=("Arial", 12), 
              bg="#B39DFF", fg="white", command=generate).pack(pady=10)
    
def show_sklad():
    sklad_window = tk.Toplevel(root)
    sklad_window.title("Склад и управление остатками")
    sklad_window.geometry("900x750")
    sklad_window.configure(bg="#E6D9FF")

    tk.Label(sklad_window, text="⚠️ Топ-5 товаров с малым остатком", 
             font=("Arial", 14, "bold"), bg="#E6D9FF", fg="#D32F2F").pack(pady=10)

    low_stock_columns = ("ID", "Название", "Остаток")
    low_stock_tree = ttk.Treeview(sklad_window, columns=low_stock_columns, show="headings", height=5)
    for col in low_stock_columns:
        low_stock_tree.heading(col, text=col)
        low_stock_tree.column(col, width=200)
    low_stock_tree.pack(padx=20, pady=5)

    def load_low_stock():
        low_stock_tree.delete(*low_stock_tree.get_children())
        # Берем 5 товаров, где остаток меньше всего (сортировка по возрастанию)
        cursor.execute("SELECT id_product, name_product, quanite_at_strogare FROM products ORDER BY quanite_at_strogare ASC LIMIT 5")
        for row in cursor.fetchall():
            low_stock_tree.insert("", "end", values=row)

    load_low_stock()

    tk.Frame(sklad_window, height=2, bd=1, relief="sunken", bg="grey").pack(fill="x", padx=50, pady=20)
    tk.Label(sklad_window, text="➕ Добавить новый товар", 
             font=("Arial", 14, "bold"), bg="#E6D9FF", fg="#5A3D8C").pack(pady=5)

    add_frame = tk.Frame(sklad_window, bg="#E6D9FF")
    add_frame.pack(pady=10)

    labels = ["Название:", "Цена:", "Категория (ID):", "Количество:"]
    entries = []
    for i, text in enumerate(labels):
        tk.Label(add_frame, text=text, bg="#E6D9FF", font=("Arial", 11)).grid(row=i, column=0, sticky="e", padx=5, pady=5)
        entry = tk.Entry(add_frame, width=30)
        entry.grid(row=i, column=1, padx=5, pady=5)
        entries.append(entry)

    ent_name, ent_price, ent_cat, ent_qty = entries

    def add_product():
        name = ent_name.get()
        try:
            price = float(ent_price.get())
            cat_id = int(ent_cat.get())
            qty = float(ent_qty.get())
            
            if not name: raise ValueError

            cursor.execute("""
                INSERT INTO products (name_product, price, id_category, quanite_at_strogare) 
                VALUES (?, ?, ?, ?)
            """, (name, price, cat_id, qty))
            
            connection.commit()
            messagebox.showinfo("Успех", f"Товар '{name}' добавлен!")
            
            for e in entries: e.delete(0, tk.END)
            load_low_stock()
            
        except ValueError:
            messagebox.showerror("Ошибка", "Проверьте правильность заполнения полей!")
        except sqlite3.Error as e:
            messagebox.showerror("Ошибка БД", f"Не удалось добавить: {e}")

    tk.Button(sklad_window, text="Сохранить товар", font=("Arial", 12, "bold"), 
              bg="#8A6CFF", fg="white", width=20, command=add_product).pack(pady=15)

    tk.Frame(sklad_window, height=2, bd=1, relief="sunken", bg="grey").pack(fill="x", padx=50, pady=20)
    tk.Label(sklad_window, text="📦 Пополнение запасов (уже существующих)", 
             font=("Arial", 14, "bold"), bg="#E6D9FF", fg="#2E7D32").pack(pady=5)

    refill_frame = tk.Frame(sklad_window, bg="#E6D9FF")
    refill_frame.pack(pady=10)

    tk.Label(refill_frame, text="ID товара:", bg="#E6D9FF").grid(row=0, column=0, padx=5)
    ent_refill_id = tk.Entry(refill_frame, width=10)
    ent_refill_id.grid(row=0, column=1, padx=5)

    tk.Label(refill_frame, text="Сколько привезли:", bg="#E6D9FF").grid(row=0, column=2, padx=5)
    ent_refill_qty = tk.Entry(refill_frame, width=10)
    ent_refill_qty.grid(row=0, column=3, padx=5)

    def refill_stock():
        try:
            p_id = int(ent_refill_id.get())
            add_qty = float(ent_refill_qty.get())
            
            cursor.execute("SELECT name_product FROM products WHERE id_product = ?", (p_id,))
            product = cursor.fetchone()

            if product:
                cursor.execute("UPDATE products SET quanite_at_strogare = quanite_at_strogare + ? WHERE id_product = ?", 
                               (add_qty, p_id))
                connection.commit()
                messagebox.showinfo("Готово", f"Запасы товара '{product[0]}' пополнены на {add_qty}")
                ent_refill_id.delete(0, tk.END)
                ent_refill_qty.delete(0, tk.END)
                load_low_stock()
            else:
                messagebox.showerror("Ошибка", "Товар с таким ID не найден!")
        except:
            messagebox.showerror("Ошибка", "Введите числа в поля ID и Количество!")

    tk.Button(sklad_window, text="Принять поставку", font=("Arial", 12, "bold"), 
              bg="#4CAF50", fg="white", width=25, command=refill_stock).pack(pady=15)
    
root = tk.Tk()
root.title("Магазин")
root.geometry("650x480")
root.configure(bg="#E6D9FF")
tk.Label(root, text="Система учёта товаров в магазине", 
         font=("Arial", 19, "bold"), bg="#E6D9FF", fg="#5A3D8C").pack(pady=50)
tk.Button(root, text="Новая продажа", font=("Arial", 14), width=28, height=3,
          bg="#B39DFF", fg="white", command=open_new_sale).pack(pady=20)
tk.Button(root, text="Отчёты", font=("Arial", 14), width=28, height=3,
          bg="#B39DFF", fg="white", command=show_reports).pack(pady=15)
tk.Button(root, text="Склад", font=("Arial", 14), width=28, height=3,
          bg="#B39DFF", fg="white", command=show_sklad).pack(pady=15)
root.mainloop()
connection.close()